#!/usr/bin/env python3
"""
학교 엑셀 + 관리자 HTML option 목록을 합쳐 schools-data.js를 생성하는 스크립트

예시:
python tools/build_school_data.py \
  --xlsx "../강원특별자치도교육청 산하 학교 현황.xlsx" \
  --html "../붙여넣은 텍스트 (1).txt" \
  --out "../assets/js/schools-data.js"
"""

from __future__ import annotations

import argparse
import json
import pathlib
import re
from typing import Dict, List

import openpyxl
from bs4 import BeautifulSoup

REGION_ORDER = ["강릉","고성","동해","삼척","속초","양구","양양","영월","원주","인제","정선","철원","춘천","태백","평창","홍천","화천","횡성"]


def clean_address(address: str | None) -> str:
    value = "" if address is None else str(address).strip()
    value = re.sub(r"\s+", " ", value)
    value = value.replace("윈주시", "원주시")
    return value


def derive_region(address: str | None) -> str:
    value = clean_address(address)
    for prefix in ("강원특별자치도 ", "강원도 "):
        if value.startswith(prefix):
            value = value[len(prefix):]
    first = value.split()[0] if value else ""
    if first.endswith(("시", "군")):
        first = first[:-1]
    if first == "윈주":
        first = "원주"
    return first


def derive_level(name: str) -> str:
    value = (name or "").strip()
    if "유치원" in value:
        return "유치원"
    if "초중학교" in value:
        return "초중"
    if "고등학교" in value:
        return "고등"
    if "중학교" in value:
        return "중등"
    if "초등학교" in value or "초병설" in value or "분교장" in value:
        return "초등"
    if "학교" in value:
        return "특수/기타"
    return "기타"


def normalize_website(value: str | None) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if not re.match(r"^https?://", text):
        text = "https://" + text
    return text


def load_school_codes(html_path: pathlib.Path) -> Dict[str, str]:
    text = html_path.read_text(encoding="utf-8")
    soup = BeautifulSoup(text, "html.parser")
    select = soup.find("select", {"name": "neisCD"})
    mapping: Dict[str, str] = {}
    if not select:
        return mapping
    for option in select.find_all("option"):
        code = (option.get("value") or "").strip()
        name = option.get_text(strip=True)
        if code and name:
            mapping[name] = code
    return mapping


def load_excel_rows(xlsx_path: pathlib.Path) -> List[dict]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    data_rows = rows[1:]

    school_codes: Dict[str, str] = {}
    return data_rows


def build_dataset(xlsx_path: pathlib.Path, html_path: pathlib.Path) -> List[dict]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    data_rows = rows[1:]
    code_map = load_school_codes(html_path)

    records: List[dict] = []
    for index, row in enumerate(data_rows, start=1):
        name, address, phone, website = row[:4]
        school_name = "" if name is None else str(name).strip()
        school_address = clean_address(address)
        school_phone = "" if phone is None else str(phone).strip()
        school_website = normalize_website(website)
        school_region = derive_region(school_address)
        school_code = code_map.get(school_name, "")
        school_id = school_code if school_code else f"SCH{index:04d}"

        records.append(
            {
                "id": school_id,
                "schoolCode": school_code,
                "name": school_name,
                "region": school_region,
                "address": school_address,
                "phone": school_phone,
                "website": school_website,
                "level": derive_level(school_name),
            }
        )

    order_index = {name: idx for idx, name in enumerate(REGION_ORDER)}
    records.sort(key=lambda item: (order_index.get(item["region"], 999), item["name"]))
    return records


def export_js(records: List[dict], out_path: pathlib.Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    content = "window.SCHOOLS_DATA = " + json.dumps(records, ensure_ascii=False, separators=(",", ":")) + ";\n"
    out_path.write_text(content, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="학교 엑셀 파일 경로")
    parser.add_argument("--html", required=True, help="관리자 HTML 텍스트 경로")
    parser.add_argument("--out", required=True, help="출력 JS 파일 경로")
    args = parser.parse_args()

    xlsx_path = pathlib.Path(args.xlsx).resolve()
    html_path = pathlib.Path(args.html).resolve()
    out_path = pathlib.Path(args.out).resolve()

    records = build_dataset(xlsx_path, html_path)
    export_js(records, out_path)

    matched = sum(1 for item in records if item["schoolCode"])
    print(f"생성 완료: {out_path}")
    print(f"- 총 학교 수: {len(records)}")
    print(f"- 코드 매칭 수: {matched}")
    print(f"- 코드 미매칭 수: {len(records) - matched}")


if __name__ == "__main__":
    main()
